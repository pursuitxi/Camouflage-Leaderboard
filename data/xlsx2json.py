import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


TASK_INFO = {
    "id": "cod",
    "name": "Camouflaged Object Detection",
}

DATASETS = [
    {"id": "camo", "name": "CAMO"},
    {"id": "cod10k", "name": "COD10K"},
    {"id": "nc4k", "name": "NC4K"},
]

METRICS = ["MAE", "WeightedFmeasure", "Smeasure", "Emeasure", "ContextMeasure"]

EXCEL_TO_METRIC = {
    "MAE": "MAE",
    "WeightedFmeasure": "WeightedFmeasure",
    "Smeasure": "Smeasure",
    "Emeasure": "Emeasure",
    "CMeasure": "ContextMeasure",
}

HIGHER_IS_BETTER = {
    "MAE": False,
    "WeightedFmeasure": True,
    "Smeasure": True,
    "Emeasure": True,
    "ContextMeasure": True,
}


def normalize_size(value):
    if value is None:
        return None
    return str(value).replace("×", "x")


def sheet_id(sheet_name):
    return sheet_name.strip().lower()


def read_dataset_headers(sheet):
    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        dataset_name = sheet.cell(1, col_idx).value
        if dataset_name:
            headers.append((str(dataset_name).strip(), col_idx))
    return headers


def read_category(sheet):
    dataset_headers = read_dataset_headers(sheet)
    models = []

    for row_idx in range(3, sheet.max_row + 1):
        method_name = sheet.cell(row_idx, 1).value
        if not method_name:
            continue

        model = {
            "name": str(method_name).strip(),
            "pubYear": sheet.cell(row_idx, 2).value,
            "size": normalize_size(sheet.cell(row_idx, 3).value),
            "backbone": sheet.cell(row_idx, 4).value,
            "paper": None,
            "code": None,
            "results": {},
        }

        for dataset_name, start_col in dataset_headers:
            dataset_id = dataset_name.strip().lower()
            model["results"][dataset_id] = {}
            for offset in range(len(METRICS)):
                excel_metric = sheet.cell(2, start_col + offset).value
                metric = EXCEL_TO_METRIC.get(excel_metric)
                value = sheet.cell(row_idx, start_col + offset).value
                if metric and isinstance(value, (int, float)):
                    model["results"][dataset_id][metric] = float(value)

        models.append(model)

    return {
        "id": sheet_id(sheet.title),
        "name": sheet.title,
        "datasets": DATASETS,
        "metrics": METRICS,
        "higherIsBetter": HIGHER_IS_BETTER,
        "defaultMetric": "ContextMeasure",
        "models": models,
    }


def build_leaderboard_json(input_xlsx):
    workbook = load_workbook(input_xlsx, data_only=True)
    categories = [read_category(sheet) for sheet in workbook.worksheets]

    return {
        "tasks": [
            {
                "id": TASK_INFO["id"],
                "name": TASK_INFO["name"],
                "categories": categories,
            }
        ]
    }


def rank_sum(models, dataset_ids, metric, higher_is_better):
    totals = {model["name"]: 0 for model in models}
    sortable_models = {model["name"]: model for model in models}

    for dataset_id in dataset_ids:
        scored = []
        for model in models:
            score = model.get("results", {}).get(dataset_id, {}).get(metric)
            if isinstance(score, (int, float)):
                scored.append((model["name"], score))

        scored.sort(key=lambda item: item[1], reverse=higher_is_better[metric])

        for index, (name, _) in enumerate(scored, start=1):
            totals[name] += index

    ranked = []
    for name, total in totals.items():
        model = dict(sortable_models[name])
        model["rankSum"] = total
        ranked.append(model)

    ranked.sort(key=lambda model: (model["rankSum"], model["name"]))
    return ranked


def main():
    parser = argparse.ArgumentParser(description="Convert benchmark.xlsx to leaderboards.json")
    parser.add_argument("--input", type=Path, default=Path(__file__).with_name("benchmark.xlsx"))
    parser.add_argument("--output", type=Path, default=Path(__file__).with_name("leaderboards.json"))
    args = parser.parse_args()

    payload = build_leaderboard_json(args.input)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Saved to {args.output}")


if __name__ == "__main__":
    main()
